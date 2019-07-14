import requests
import datetime
import xmltodict
import os
import json
import uuid
import xlsxwriter
from multiprocessing import process
import openpyxl
from bottle import abort, request, static_file, run, route
import sys
from clinical_trials import Trials
import zipfile
reload(sys)
sys.setdefaultencoding("utf-8")

temp_path = os.path.dirname(os.path.abspath(__file__)) + os.path.sep + "tmp"
if not os.path.exists(temp_path):
    os.makedirs(temp_path)


SHEET_LIMIT = 50000
MAX_SHEETS_PER_XLS = 7
MY_API_KEY = "6f63b0b5ec41afd50bed862a0d61ff0ae709"
PUBMED_SEARCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db=pubmed&api_key=6f63b0b5ec41afd50bed862a0d61ff0ae709&term={}"
PUBMED_DATE_QUERY = '+AND+("{}"[PDat] : "{}"[PDat])'
#PUBMED_DOWNLOAD_CSV =  "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?db=pubmed&id={}&rettype=fasta&retmode=xml&api_key=6f63b0b5ec41afd50bed862a0d61ff0ae709"
PUBMED_DOWNLOAD_CSV =  "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?db=pubmed&rettype=fasta&retmode=xml&api_key=6f63b0b5ec41afd50bed862a0d61ff0ae709"
#Date_format:YYYY/MM/DD
pubmed_headers = ["GM Universal Code", "Full Name", "Author Match","Authorship_Position", "Publication_Type", "Mesh_Headings","Title","URL", "Query Used",
          "Description","Details","ShortDetails", "Affiliation","Resource","Type","Identifiers","Db","EntrezUID","Properties", "Author_Count", "Abstract_Text"]

trails_headers = ['GM Universal Code', 'Full Name', 'NCT ID', 'URL', 'Verification Status', 'Query Used', 'Trial Name', "Trial Type",'Trial Phase' , 'Overall Status',
 'Start Date', 'End Date', 'Conditions', 'Interventions', 'Matched Associate','Role', 'Facility', 'Region', 'Other Associates', 'Organizations', 'Lead Sponsor(s)']


def zipper(zip_file_name, files):
    zip_file_name = '{}{}{}.zip'.format(temp_path, os.path.sep, zip_file_name)
    print("Zipping {} xlsx files to {}".format(len(files), zip_file_name))
    with zipfile.ZipFile(zip_file_name,'w') as zip_:
        for file_ in files:
            zip_.write(temp_path + os.path.sep + file_)
    print("file:{}".format(zip_file_name))
    return os.path.basename(zip_file_name)
        

def create_xlsx(data=None, data_list=[], local=False,headers=pubmed_headers,sheet_limit=SHEET_LIMIT):
    main_file_name = str(uuid.uuid4())+".xlsx"
    sheet_number = 1
    workbook = xlsxwriter.Workbook(temp_path + os.path.sep + main_file_name)
    worksheet = workbook.add_worksheet(name="Sheet{}".format(sheet_number))
    row = 0
    col = 0
    total_files = [main_file_name]
    for header in headers:
        worksheet.write(row, col, header)
        col += 1

    if not local:
        for col_data in data:
            row += 1
            col = 0
            for header in headers:
                #print("header--> {}:data-->{}:type--->{}".format(header,col_data[header],type(col_data[header])))
                worksheet.write(row, col, col_data[header])
                col += 1

            if sheet_number < MAX_SHEETS_PER_XLS:
                if row >= sheet_limit:
                    row = 0
                    col = 0
                    sheet_number += 1
                    worksheet = workbook.add_worksheet(name="Sheet{}".format(sheet_number))
                    for header in headers:
                        worksheet.write(row, col, header)
                        col += 1
            else:
                if row < sheet_limit:
                    continue
                workbook.close() 
                file_name = "{}_part-{}.xlsx".format(os.path.splitext(main_file_name)[0], len(total_files))
                sheet_number = 1
                workbook = xlsxwriter.Workbook(temp_path + os.path.sep + file_name)
                worksheet = workbook.add_worksheet(name="Sheet{}".format(sheet_number))
                row = 0
                col = 0
                for header in headers:
                    worksheet.write(row, col, header)
                    col += 1
                total_files.append(file_name)
                
    else:
        for data in data_list:
            for col_data in data:
                row += 1
                col = 0
                for header in headers:
                    #print("header--> {}:data-->{}:type--->{}".format(header,col_data[header],type(col_data[header])))
                    worksheet.write(row, col, col_data[header])
                    col += 1

                if sheet_number < MAX_SHEETS_PER_XLS:
                    if row >= sheet_limit:
                        row = 0
                        col = 0
                        sheet_number += 1
                        worksheet = workbook.add_worksheet(name="Sheet{}".format(sheet_number))
                        for header in headers:
                            worksheet.write(row, col, header)
                            col += 1
                else:
                    if row < sheet_limit:
                        continue
                    workbook.close() 
                    file_name = "{}_part-{}.xlsx".format(os.path.splitext(main_file_name)[0], len(total_files))
                    sheet_number = 1
                    workbook = xlsxwriter.Workbook(temp_path + os.path.sep + file_name)
                    worksheet = workbook.add_worksheet(name="Sheet{}".format(sheet_number))
                    row = 0
                    col = 0
                    for header in headers:
                        worksheet.write(row, col, header)
                        col += 1
                    total_files.append(file_name)  

    workbook.close()      

    if len(total_files) == 1:
        return main_file_name
    else:
        return zipper(os.path.splitext(main_file_name)[0], total_files)

def xml_to_json(xml):
    ''' This API converts xml data to json
        parameters:
        ---------------
        xml (str) : Complete xml data read from website or file

        return:
        ---------------
        json(dict) : returns valid dictionary
    '''
    json_string = json.dumps(xmltodict.parse(xml))
    json_data = json.loads(json_string)
    return json.dumps(json_data)

def get_description(auther_list):
    desc = ""
    i = 0
    if type(auther_list) == dict:
        auther_list = [auther_list]
    for author in auther_list:
        i += 1
        if author.get("CollectiveName",None):
            desc += author["CollectiveName"]
        else:
            if author.get("LastName", "") is None:
                author["LastName"] = ""
            if author.get("ForeName", "") is None:
                author["ForeName"] = ""
            if author.get("Initials", "") is None:
                author["Initials"] = ""
            desc += author.get("LastName","") + " " + author.get("Initials","")
        if len(auther_list) == i:
            desc += "."
        else:
            desc += ", "
    return desc 

def get_affiliation_details(auther_name, auther_list):
    if type(auther_list) == dict:
        auther_list = [auther_list]
    name_list = auther_name.replace(",","").lower().split(" ")
    for author in auther_list:
        if author.get("AffiliationInfo"):
            if type(author["AffiliationInfo"]) == dict:
                author["AffiliationInfo"] = [author["AffiliationInfo"]]
                data = ""
                i = 0
                if author.get("LastName", "") is None:
                    author["LastName"] = ""
                if author.get("ForeName", "") is None:
                    author["ForeName"] = ""
                if author.get("LastName", "").lower() in name_list or author.get("ForeName", "").lower() in name_list:
                    for aff_info in author["AffiliationInfo"]:
                        data += aff_info["Affiliation"]
                        i = i+ 1
                        if len(author["AffiliationInfo"]) != i:
                            data += ", "
                        else:
                            data += "."
 
                    return data
    return ""

def get_journal_issue_details(journal_issue):
    data = "{} {} {};".format(journal_issue["PubDate"].get("Year",""), journal_issue["PubDate"].get("Month",""),journal_issue["PubDate"].get("Day",""))
    if journal_issue.get("Volume"):
        data += " {}({}).".format(journal_issue["Volume"],journal_issue.get("Issue",0)) 
    return data

def get_elocation_details(elocationid):
    if type(elocationid) == dict:
        elocationid = [elocationid]
    data = ""
    for value in elocationid:
        data += "{}: {}. ".format(value["@EIdType"],value["#text"])
    return data

def get_details(journal, elocationid):
    data = "{}. {} {}".format(journal.get("ISOAbbreviation",""),get_journal_issue_details(journal["JournalIssue"]),get_elocation_details(elocationid))
    return data

def get_short_details(journal):
    data = "{}. {}".format(journal.get("ISOAbbreviation", ""), journal["JournalIssue"]["PubDate"].get("Year",""))
    return data

def get_create_date(pub_dates):
    if type(pub_dates) == dict:
        pub_dates = [pub_dates]
    for date in pub_dates:
        if date["@PubStatus"] in ["pubmed", "medline"]:
            data = "{}/{}/{}".format(date.get("Year",""), date["Month"], date["Day"])
            return data
    return ""
    
def get_first_author(auther_list):
    if type(auther_list) == dict:
        auther_list = [auther_list]
    data = ""
    for author in auther_list:
        try:
            if author.get("LastName", "") is None:
                author["LastName"] = ""
            if author.get("ForeName", "") is None:
                author["ForeName"] = ""
            if author.get("Initials", "") is None:
                author["Initials"] = ""
            data = author.get("LastName","") + " " + author.get("Initials","")
            return data
        except KeyError:
            if author.get("CollectiveName",None):
                data = author["CollectiveName"]
    return data    
        
def get_full_name(name, auther_list):
    if type(auther_list) == dict:
        auther_list = [auther_list]
    name_list = name.replace(",","").lower().split(" ")
    for author in auther_list:
        if author.get("LastName", "") is None:
            author["LastName"] = ""
        if author.get("ForeName", "") is None:
            author["ForeName"] = ""
        if author.get("LastName", "").lower() in name_list or author.get("ForeName", "").lower() in name_list:
            return author.get("LastName", "") + ", " + author.get("ForeName","")
    return name

def get_author_position(name, auther_list):
    if type(auther_list) == dict:
        auther_list = [auther_list]
    name_list = name.replace(",","").lower().split(" ")
    for author in auther_list:
        if author.get("LastName", "") is None:
            author["LastName"] = ""
        if author.get("ForeName", "") is None:
            author["ForeName"] = ""
        if author.get("LastName", ""):
            if author.get("LastName", "").lower() in name_list:
                if author.get("ForeName", ""):
                    forename = author.get("ForeName", "").split(" ")[0]
                    if forename.lower() in name_list:
                        return auther_list.index(author) + 1
                else:
                    return auther_list.index(author) + 1
    return 0


def get_publication_type(publication_type_list):
    data = []
    if type(publication_type_list) == dict:
        publication_type_list = [publication_type_list]
    for type_ in publication_type_list:
        data.append(type_["#text"])
    return ";".join(data)

def get_mesh_headings(mesh_heading_list):
    data = []
    if type(mesh_heading_list) == dict:
        mesh_heading_list = [mesh_heading_list]
    for heading in mesh_heading_list:
        if heading.get("DescriptorName",""):
            data.append(heading["DescriptorName"].get("#text",""))
    return ";".join(data)


def get_properties(pub_dates,auther_list):
    data = "create date:{} | first author:{}".format(get_create_date(pub_dates), get_first_author(auther_list))
    return data 

@route('/upload', method='POST')
def do_upload():

    upload     = request.files.get('upload')
    name, ext = os.path.splitext(upload.filename)
    
    if ext in ('.png','.jpg','.jpeg'):
        return 'File extension not allowed.'

    upload.save(temp_path) # appends upload.filename automatically
    
    from_date = request.forms.get('from_date') if request.forms.get('from_date') else None
    to_date = request.forms.get('to_date') if request.forms.get('to_date') else None
    search_type = request.forms.get("search_type") if request.forms.get('search_type') else None
    sheet_len = int(request.forms.get("sheet_len")) if request.forms.get('sheet_len') else SHEET_LIMIT 
    
    header_ = pubmed_headers
    if search_type == "Clinical Trails":
        header_ = trails_headers
    if from_date and to_date:
        if datetime.datetime.strptime(from_date, "%Y-%m-%d") > datetime.datetime.strptime(to_date, "%Y-%m-%d"):
            return '<html><script>alert("from date should be lesser than to date");</script><html>'

    try:
        xlsx_file_path = os.path.join(temp_path, upload.filename)
        xlsx_data = []
        wb_obj = openpyxl.load_workbook(xlsx_file_path)
        sheet_obj = wb_obj.active
        for i in range(1, sheet_obj.max_row+1):
            row_data = {}
            for j in range(1, sheet_obj.max_column+1):
                row_data[sheet_obj.cell(row = 1, column = j).value] = sheet_obj.cell(row = i+1, column = j).value
            xlsx_data.append(row_data)
        
        os.remove(xlsx_file_path)
        ids_return_data = {"ids_info":{},"count":0}

        xlsx_data_list = []

        for column_data in xlsx_data:
            print("\n\n{} of {}\n\n".format(xlsx_data.index(column_data)+1, len(xlsx_data)))
            name = column_data["Full_Name"] if column_data.get("Full_Name") else None
            uid = column_data["KOL_ID"] if column_data.get("KOL_ID") else None
            firstname = column_data["First_Name"] if column_data.get("First_Name") else None
            initial = column_data["Middle_Name"] if column_data.get("Middle_Name") else None
            lastname = column_data["Last_Name"] if column_data.get("Last_Name") else None
            if name is not None:
                try:
                    search_data = search_citations(name=name, search_type=search_type,sheet_len=sheet_len,initial=initial, lastname=lastname, firstname=firstname, universal_id=uid, local_searh=True, from_date=from_date, to_date=to_date, records_per_page=4000)
                except Exception as ex:
                    if xlsx_data_list:
                        break
                    raise Exception(str(ex))

                if search_type == "Clinical Trails":
                    if search_data is None:
                        continue
                    xlsx_data_list.append(search_data)
                    ids_return_data = {"ids_info":{},"count":1}
                    continue

                if search_data["count"] != 0:
                    ids_return_data["ids_info"].update(search_data["ids_info"])                
                    ids_return_data["count"] += len(search_data["ids_info"].keys())
                    xlsx_data_list.append(download_csv(search_data, local=True))

        if ids_return_data["count"] != 0:
            
            file_path = create_xlsx(data_list=xlsx_data_list, local=True, headers=header_, sheet_limit=sheet_len)
            return static_file(file_path, temp_path, download=file_path)
    except Exception as ex:
        print("Exception in upload:{}".format(ex))
        abort(500, "Exception occurred: {}".format(ex))        


@route("/search", method='POST')
def search_citations(name=None, search_type="Pubmed",initial=None, lastname=None, firstname=None, universal_id=None,from_date=None, to_date=None,  records_per_page="400", local_searh=False,sheet_len=SHEET_LIMIT):
    try:
        if not local_searh:
            name = request.forms.get('Name')  if request.forms.get('Name') else None
            records_per_page = request.forms.get('records') if request.forms.get('records') else "420"
            universal_id = request.forms.get('Uid') if request.forms.get('Uid') else str(uuid.uuid4())
            from_date = request.forms.get('from_date')  if request.forms.get('from_date') else None
            to_date = request.forms.get('to_date')  if request.forms.get('to_date') else None
            initial = request.forms.get('Initial')  if request.forms.get('Initial') else None
            lastname = request.forms.get('Lastname')  if request.forms.get('Lastname') else None
            firstname = request.forms.get("FirstName") if request.forms.get('FirstName') else None
            search_type = request.forms.get("search_type") if request.forms.get('search_type') else None
            sheet_len = int(request.forms.get("sheet_len")) if request.forms.get('sheet_len') else SHEET_LIMIT
        
        if search_type == "Clinical Trails":
            file_ = clinical_trails(name, universal_id, lastname=lastname, initial=initial, firstname=firstname,local=local_searh,sheet_limit=sheet_len)
            if local_searh:
                return file_
            else:
                if file_ is not None:
                    return static_file(file_, temp_path, download=file_)
                return "No Data found"

        url = None
        if not name:
            raise Exception("Name is madatory field")

        search_name = "" + name + "[Full Author Name]"
        if search_name:
        
            if firstname and initial and lastname:
                search_name += '+OR+"{}{} {}"[Full Author Name]'.format(firstname[0], initial[0],lastname)
                search_name += '+OR+"{}, {}{}"[Full Author Name]'.format(lastname, firstname[0], initial[0])
                search_name += '+OR+"{} {}{}"[Author]'.format(lastname, firstname[0], initial[0])
                search_name += '+OR+"{}, {} {}"[Full Author Name]'.format(lastname, firstname, initial)
                search_name += '+OR+"{}, {} {}"[Full Author Name]'.format(lastname, firstname, initial[0])
                search_name += '+OR+"{} {} {}"[Author]'.format(initial, lastname, firstname)
                 
            if lastname and firstname:
                search_name += '+OR+"{}, {}"[Full Author Name]'.format(lastname, firstname)
                search_name += '+OR+"{} {}"[Author]'.format(firstname, lastname)
                search_name += '+OR+"{}, {}"[Full Author Name]'.format(lastname, firstname[0])
                search_name += '+OR+"{} {}"[Author]'.format(lastname, firstname[0])


            url = PUBMED_SEARCH_URL.format(search_name)

        if from_date and to_date:
            if datetime.datetime.strptime(from_date, "%Y-%m-%d") > datetime.datetime.strptime(to_date, "%Y-%m-%d"):
                return '<html><script>alert("from date should be lesser than to date");</script><html>'

            url = url + PUBMED_DATE_QUERY.format(from_date,to_date)
        
        query_url = url.split("&term=")[-1]

        if not url:
            raise Exception("Unable to build search url")
        url += "&retmax={}".format(records_per_page)

        print("pubmed search url: {}".format(url))
        r = requests.post(url=url,headers= {"Content-Type": "application/xml","accept": "application/xml"})
        if r.status_code == 200:
            json_data = json.loads(xml_to_json(r.text))
        else:
            raise Exception("Error occured while comunicating with pubmed{}".format(r.status_code))
        
        if int(json_data["eSearchResult"]["Count"]) <= 0:
            return {"ids_info":{},"count":0 }


        id_list = json_data["eSearchResult"]["IdList"].get("Id",None) 
      
        if type(id_list) in [str, int, unicode]:
            id_list = [id_list]

        return_data = {"ids_info":{}}
        for _id in id_list:
            return_data["ids_info"][_id] = {"name": name,"query":query_url,"univeral_id":universal_id}

        return_data["count"] = len(return_data["ids_info"].keys())

        print("Total records found ------------->{}".format(return_data["count"]))
        if return_data["count"] != 0:

            if not local_searh:
                file_path = download_csv(query_data=return_data, local=local_searh,sheet_limit=sheet_len)
                return static_file(file_path, temp_path, download=file_path)
            else:
                return return_data

    except Exception as ex:
        print("search_citations:Exception occurred: {}".format(ex))
        abort(500, "Exception occurred: {}".format(ex))


@route("/pubmed/download")
def download_csv(query_data=None, local=False,sheet_limit=SHEET_LIMIT):
    try:
        xlsx_data = []
        # if not local:
        #     query_data = json.loads(request.query.data) if request.query.data else None
    
        if not query_data:
            raise Exception("query_data is madatory field")
        if not query_data.get("count") or not query_data.get("ids_info"):
            raise Exception("Invalid data provided")

        if len(query_data["ids_info"].keys()) == 1 :
            ids = query_data["ids_info"].keys()[0]
        else:
            ids = ",".join(query_data["ids_info"].keys())

        url = PUBMED_DOWNLOAD_CSV
        print("pubmed download csv url: {}".format(url))
        r = requests.post(url=url, data="id={}".format(ids),headers= {"accept": "application/xml"})
        if r.status_code == 200:
            json_data = json.loads(xml_to_json(r.text))
            if type(json_data["PubmedArticleSet"]["PubmedArticle"]) == dict:
                json_data["PubmedArticleSet"]["PubmedArticle"] = [json_data["PubmedArticleSet"]["PubmedArticle"]]
            
            data_count = 1
            for data in json_data["PubmedArticleSet"]["PubmedArticle"]:
                medline_data = data["MedlineCitation"]
                article_data = medline_data["Article"]
                publication_date = data["PubmedData"]
                mesh_heading_data = medline_data.get("MeshHeadingList",{})
                form_data = {}
                print("Name---->{}, data found-->({}/{})".format( query_data["ids_info"][medline_data["PMID"]["#text"]]["name"], data_count,len(query_data["ids_info"].keys())))
                data_count += 1
                if type(article_data["ArticleTitle"]) == dict:
                    form_data["Title"] = article_data["ArticleTitle"]["#text"]
                else:
                    form_data["Title"] = article_data["ArticleTitle"]
                form_data["URL"] = "https://www.ncbi.nlm.nih.gov/pubmed/" + medline_data["PMID"]["#text"]
                form_data["Description"] = get_description(article_data["AuthorList"]["Author"])
                if article_data.get("ELocationID"):
                    form_data["Details"] = get_details(article_data["Journal"], article_data["ELocationID"])
                else:
                    
                    form_data["Details"] = "{}. {}".format(article_data["Journal"].get("ISOAbbreviation",""),get_journal_issue_details(article_data["Journal"]["JournalIssue"]))

                form_data["GM Universal Code"] = query_data["ids_info"][medline_data["PMID"]["#text"]]["univeral_id"]
                form_data["Affiliation"] = get_affiliation_details(query_data["ids_info"][medline_data["PMID"]["#text"]]["name"], article_data["AuthorList"]["Author"])
                form_data["Query Used"] = query_data["ids_info"][medline_data["PMID"]["#text"]]["query"]
                form_data["ShortDetails"] = get_short_details(article_data["Journal"])
                form_data["Resource"] = "PubMed"
                form_data["Type"] = "citation"
                form_data["Full Name"] =   query_data["ids_info"][medline_data["PMID"]["#text"]]["name"]    #get_full_name(query_data["ids_info"][medline_data["PMID"]["#text"]]["name"], article_data["AuthorList"]["Author"])
                form_data["Identifiers"] = "PMID:" + medline_data["PMID"]["#text"]
                form_data["Db"] = "pubmed"
                if article_data.get("PublicationTypeList"):
                    form_data["Publication_Type"] = get_publication_type(article_data["PublicationTypeList"].get("PublicationType",[]))
                else:
                    form_data["Publication_Type"] = ""
                
                if mesh_heading_data.get("MeshHeading"):
                    form_data["Mesh_Headings"] = get_mesh_headings(mesh_heading_data.get("MeshHeading",[]))
                else:
                    form_data["Mesh_Headings"] = ""
                
                form_data["Authorship_Position"] = get_author_position(query_data["ids_info"][medline_data["PMID"]["#text"]]["name"], article_data["AuthorList"]["Author"])
                form_data["Author_Count"] = len(article_data["AuthorList"]["Author"])
                form_data["Abstract_Text"] = ""
                if article_data.get("Abstract",None):
                    if article_data["Abstract"].get("AbstractText",""):
                        if type(article_data["Abstract"]["AbstractText"]) == list:
                            for text in article_data["Abstract"]["AbstractText"]:
                                if text:
                                    if type(text) == dict:
                                        form_data["Abstract_Text"] += text.get("#text","")
                                    else:
                                        form_data["Abstract_Text"] += text   
                                form_data["Abstract_Text"] += ";"

                        elif type(article_data["Abstract"]["AbstractText"]) == dict:
                            form_data["Abstract_Text"] = article_data["Abstract"]["AbstractText"].get("#text")
                        else:
                            form_data["Abstract_Text"] = article_data["Abstract"]["AbstractText"]

                form_data["EntrezUID"] = medline_data["PMID"]["#text"]
                form_data["Author Match"] = get_full_name(query_data["ids_info"][medline_data["PMID"]["#text"]]["name"], article_data["AuthorList"]["Author"])
                form_data["Properties"] = get_properties(publication_date["History"]["PubMedPubDate"], article_data["AuthorList"]["Author"])
                xlsx_data.append(form_data)
        else:
            raise Exception("Error occured while comunicating with pubmed, status_code({}),text({})".format(r.status_code,r.text))

        if not xlsx_data:
            raise Exception("Error occured while comunicating with pubmed")
        
        if local:
            return xlsx_data

        file_name = create_xlsx(data=xlsx_data, local=False, sheet_limit=sheet_limit)

        return file_name
    except Exception as ex:
        print("download_csv:Exception occurred: {}".format(ex))
        if local:
            if xlsx_data:
                return xlsx_data
        abort(500, "Exception occurred: {}".format(ex))


def get_interventions(interventions_list):
    data = ""
    i = 0
    if type(interventions_list) == dict:
        interventions_list = [interventions_list]
    for interventions in interventions_list:
        data += interventions["intervention_name"]
        i += 1
        if i != len(interventions_list):
            data += " | "

    return data

def get_role(overall_list):
    if type(overall_list) == dict:
        overall_list = [overall_list]
    data = ""
    i = 0
    for v in overall_list:
        data += v.get("role","")
        i += 1
        if i != len(overall_list):
            data += " | "
    return data

def get_facilities(facilities, name, last_name=None):
    if type(facilities) == dict:
        facilities = [facilities]
    search_name = name
    if last_name:
        search_name = [last_name]
    else:
        search_name = name.split(" ")


    for facility in facilities:
        if not facility["facility"].get("name"):
            continue
        for name_ in search_name:
            if name_ in facility["facility"]["name"]:
                return facility["facility"]["name"], "{}_{}_{}".format(facility["facility"]["address"].get("country"),facility["facility"]["address"].get("city"), facility["facility"]["address"].get("zip"))
        
    return "", ""

def get_other_associates(facilities):
    if type(facilities) == dict:
        facilities = [facilities]
    data = ""
    for facility in facilities:
        if not facility["facility"].get("name"):
            continue
        data += facility["facility"]["name"] + "|"
    return data

def get_sponsers(sponsers_list):
    if type(sponsers_list) == dict:
        sponsers_list = [sponsers_list]
    data = ""
    for sponser in sponsers_list:
        data += sponser["lead_sponsor"]["agency"]
    return data

def get_start_date(date_):
    if type(date_) == dict:
        return date_["#text"]
    else:
        return date_

def get_matched_associate(overall_official_list, name, lastname):
    if type(overall_official_list) == dict:
        overall_official_list = [overall_official_list]
    search_name_list = []
    if lastname:
        lastname = lastname.lower()
        lastname = lastname.replace(","," ")
        lastname_list = lastname.split(" ")
        try:
            lastname_list.remove(" ")
        except:
            pass
        search_name_list = lastname_list
    else:
        name = name.lower()
        name = name.replace(","," ")
        name_list = name.split(" ")
        try:
            name_list.remove(" ")
        except:
            pass
        search_name_list = name_list

    for official in overall_official_list:
        if official.get("last_name"):
            official_last_name = official["last_name"].lower()
            official_last_name = official_last_name.replace(","," ")
            official_last_name_list = official_last_name.split(" ")
            for name in official_last_name_list:
                if name in search_name_list:
                    return official["last_name"]
    return ""


def clinical_trails(name, _uuid, lastname=None, initial=None, firstname=None, local=False,sheet_limit=SHEET_LIMIT):
    print("name:{},lastname:{}".format(name,lastname))
    try:
        t = Trials()
        zip_folder_path = os.path.join(temp_path,str(uuid.uuid4()))
        if not os.path.exists(zip_folder_path):
            os.makedirs(zip_folder_path)
        file_path = os.path.join(zip_folder_path,str(uuid.uuid4())+ ".zip")
        zip_file_ = open(file_path,"wb")
        if lastname:
            zip_data = t.download(search_term=lastname)
        else:
            zip_data = t.download(search_term=name)
        zip_file_.write(zip_data)
        zip_file_.close()
        import zipfile
        try:
            zip_ref = zipfile.ZipFile(file_path, 'r')
            zip_ref.extractall(zip_folder_path)
            zip_ref.close()
        except Exception as ex:
            os.remove(file_path)
            os.rmdir(zip_folder_path)
            if local:
                return []
            return None

        os.remove(file_path)
        files = os.listdir(zip_folder_path)
        print("Search results--->{}".format(len(files)))
        clinical_trails_data = []
        for file_ in files:
            xml_file = os.path.join(zip_folder_path, file_)
            xml_file_obj = open(xml_file,"r")
            xml_data = xml_file_obj.read()
            json_data = json.loads(xml_to_json(xml_data))
            clinical_trails_data.append(json_data)
            xml_file_obj.close()
            os.remove(xml_file)
        os.rmdir(zip_folder_path)

        xlsx_data = []
        for data in clinical_trails_data:
            clinical_study = data["clinical_study"]
            form_data = {}
            form_data["GM Universal Code"] = _uuid
            form_data["Full Name"] = name
            form_data["NCT ID"] = clinical_study["id_info"]["nct_id"]
            form_data["URL"] = clinical_study["required_header"]["url"]
            form_data["Verification Status"] = "Probable"
            if lastname:
                form_data["Query Used"] = lastname
            else:
                form_data["Query Used"] = name
            form_data["Trial Name"] = clinical_study["brief_title"]     
            form_data["Trial Phase"] = clinical_study.get("phase","N/A")

            form_data["Overall Status"] = clinical_study["overall_status"]
            form_data["Start Date"] = get_start_date(clinical_study["start_date"]) if clinical_study.get("start_date") else ""

            if clinical_study.get("completion_date"):

                form_data["End Date"] = clinical_study["completion_date"]["#text"] if type(clinical_study["completion_date"]) == dict else clinical_study["completion_date"]
            elif clinical_study.get("primary_completion_date"):
                form_data["End Date"] = clinical_study["primary_completion_date"]["#text"]
            else:
                form_data["End Date"] = "N/A"
            form_data["Conditions"] = " | ".join(clinical_study["condition"]) if type(clinical_study["condition"]) == list else clinical_study["condition"]
            if clinical_study.get("intervention"):
                form_data["Interventions"] = get_interventions(clinical_study["intervention"])
            else:
                form_data["Interventions"] = ""
            if clinical_study.get("overall_official"):
                form_data["Role"] =  get_role(clinical_study["overall_official"])
            else:
                form_data["Role"] = ""
            form_data["Associate Type"] = "N/A"
            if clinical_study.get("location"):
                form_data['Facility'],form_data['Region'] = get_facilities(clinical_study["location"], name,lastname)
                form_data["Other Associates"] = get_other_associates(clinical_study["location"])
            else:
                form_data['Facility'] = ""
                form_data["Other Associates"] = ""
                form_data['Region'] = ""
            form_data["Trial Type"] = clinical_study.get("study_type","")
            form_data["Matched Associate"] = get_matched_associate(clinical_study.get("overall_official",[]), name, lastname)
            form_data["Organizations"] = clinical_study["source"]
            form_data["Lead Sponsor(s)"] = get_sponsers(clinical_study["sponsors"])
            xlsx_data.append(form_data)
        
        if not xlsx_data:
            raise Exception("Error occured while comunicating with clinical Trials")
        if local:
            return xlsx_data

        file_name = create_xlsx(data=xlsx_data, local=False, headers=trails_headers,sheet_limit=sheet_limit)

        return file_name

    except Exception as ex:
        print("clinical_trails exception occurred:{}".format(ex))
        if local:
            if xlsx_data:
                return xlsx_data
        raise Exception(str(ex))
        
@route("/clear_tmp",method="POST")
def clear_tmp():
    files = os.listdir(temp_path)
    for file_ in files:
        os.remove(file_)    

@route("/css/<css_file>")
def serve_css(css_file):
    return static_file(css_file, os.path.dirname(os.path.abspath(__file__))+os.path.sep+"css")

@route("/")
def serve_web():
    return static_file("index.html", os.path.dirname(os.path.abspath(__file__)))

if __name__ == "__main__":
    run(host="0.0.0.0",port=8090)
